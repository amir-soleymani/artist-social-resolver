import pandas as pd
import argparse
import os
import re
import time
from urllib.parse import urlparse, quote_plus
import requests
from typing import List, Dict, Optional
import difflib
from openpyxl import Workbook


MUSICBRAINZ_BASE = "https://musicbrainz.org/ws/2"
MB_HEADERS = {
    "User-Agent": "ArtistSocialMediaResolver/1.0 (local-script)"
}
PLATFORMS = ["Spotify", "Instagram", "apple music", "soundcloud", "youtube", "tiktok", "facebook"]
INVALID_SHEET_CHARS = r'[:\\\/\?\*\[\]]'
MAX_SHEET_NAME_LEN = 31
    
def normalize_whitespace(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()

def normalize_spotify_link(raw: str) -> str:
    """Extracts Spotify ID and returns the canonical URL format."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""

    # This pattern matches 'spotify:artist:ID' or 'open.spotify.com/artist/ID'
    m = re.search(r"artist/([A-Za-z0-9]{22})|artist:([A-Za-z0-9]{22})", s)
    if m:
        # Use whichever group (1 or 2) captured the 22-character ID
        artist_id = m.group(1) or m.group(2)
        return f"https://open.spotify.com/artist/{artist_id}"

    return s

def normalize_instagram_id(raw: str) -> str:
    """Turns input into a clean Instagram handle (no @, no URL parts)."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""

    # Remove hidden characters and URL parameters
    s = s.replace("\ufeff", "")
    s = s.split("?")[0].split("#")[0].strip()
    s = s.strip("/")

    if s.startswith("@"):
        s = s[1:]

    # If it's a full URL, take the last path segment (the handle)
    m = re.search(r"instagram\.com/([^/]+)", s)
    if m:
        s = m.group(1)

    return s

def search_duckduckgo_html(query: str, num: int = 5) -> List[Dict]:
    """
    Performs a search on DuckDuckGo and scrapes the HTML for links and titles.
    """
    # Use quote_plus to make the name safe for a URL
    ddg_url = f"https://duckduckgo.com/html/?q={quote_plus(query)}"
    headers = {"User-Agent": "Mozilla/5.0"}

    try: 
        r = requests.get(ddg_url, headers=headers, timeout=15)
        r.raise_for_status()
        html = r.text

        results = []
        
        # This Regex finds result links and titles in the DuckDuckGo HTML
        pattern = re.compile(r'<a[^>]*class="result__a"[^>]*href="([^"]+)"[^>]*>(.*?)</a>', flags=re.S)

        for m in pattern.finditer(html):
            link = m.group(1)
            title = re.sub(r"<.*?>", "", m.group(2)).strip()
            results.append({"link": link, "title": title})
            if len(results) >= num:
                break
        return results
    except Exception as e:
        print(f"Search failed for {query}: {e}")
        return []

def domain_of (url: str) -> str:
    """Extracts the domain from a URL to help categorize links."""
    try:
        return (urlparse(url).netloc or "").lower().replace("www.", "")
    except Exception:
        return ""
    
def musicbrainz_lookup_artist(artist_name: str, spotify_link: str) -> dict:
    if not artist_name:
        return {}
    
    try:
        q = quote_plus(f'artist:"{artist_name}"')
        url = f"{MUSICBRAINZ_BASE}/artist?query={q}&fmt=json&limit=8"
        r = requests.get(url, headers=MB_HEADERS, timeout=25)
        r.raise_for_status()
        data = r.json()
    except Exception:
        return {}
    
    artists = data.get("artists") or []
    if not artists:
        return {}
    
    # For now, we take the top match from MusicBrainz
    candidate = artists[0]
    mbid = candidate.get("id")

    try:
        detail_url = f"{MUSICBRAINZ_BASE}/artist/{mbid}?inc=url-rels+isnis&fmt=json"
        dr = requests.get(detail_url, headers=MB_HEADERS, timeout=25)
        dr.raise_for_status()
        detail = dr.json()
    except Exception:
        return {}
    
    links = {}
    for rel in detail.get("relations") or []:
        rel_url = ((rel.get("url") or {}).get("resource")) or ""
        if not rel_url:
            continue
        d = domain_of(rel_url)
        if "spotify.com" in d:
            links["spotify"] = rel_url
        elif "instagram.com" in d:
            links["instagram"] = rel_url
        # ("I'll add other platforms later if needed")

    isni_list = detail.get("isnis") or []
    return {
        "mbid": mbid,
        "name": detail.get("name") or artist_name,
        "isni": isni_list[0] if isni_list else "",
        "links": links,
    }

def search_duckduckgo_json(query: str, num: int = 8) -> List[Dict]:
    """
    Best-effort search without and API key using DuckDuckGo's lightweight endpoint, 
    returns a list of dicts with keys: link, title, snippet.
    """
    headers = {"User-Agent": "Mozilla/5.0"}
    # Step A: get the VQD token from the main page
    try: 
        html = requests.get(f"https://duckduckgo.com/?q={quote_plus(query)}", headers = headers, timeout = 30).text
        m = re.search(r'vqd=([\'"]?)([0-9-]+)\1', html)
        if not m:
            m = re.search(r"vqd=([0-9-]+)&", html)
        if not m:
            return []
        vqd = m.group(1)

        # Step B: use the token to get JSON results
        params = {
            "q" : query,
            "l": "us-en",
            "p": "1",
            "s": "0",
            "o": "json", 
            "vqd": vqd,
        }
        r = requests.get("https://duckduckgo.com/i.js", params = params, headers = headers, timeout = 30)
        r.raise_for_status()
        data = r.json()

        results = []
        for item in (data.get("results") or []):
            link = item.get("url") or ""
            title = item.get("title") or item.get("t") or ""
            snippet = item.get("description") or item.get("a") or ""
            if link:
                results.append({"link": link, "title": title, "snippet": snippet})
            if len(results) >= num:
                break
        return results 
    except Exception: 
        return []
    
    
def name_similarity(a: str, b: str) -> float:
    """Calculates name similarity (0...1) using token overlap and sequence matching."""
    a = (a or "").lower().strip()
    b = (b or "").lower().strip()
    if not a or not b:
        return 0.0

    # SequenceMatcher handles typos and partial matches
    ratio = difflib.SequenceMatcher(None, a, b).ratio()
    return ratio   

def pick_best_link(
        artist_name: str,
        platform: str,
        candidates: List[Dict],
) -> str: 
    """
    candidates: list of dicts with 'link' and 'title'. 
    Chooses best by name similarity and platform match.
    """
    if not candidates: 
        return ""
    
    best_link = ""
    max_score = -1.0

    for item in candidates: 
        url = (item.get("link") or "").lower()
        title = (item.get("title") or "").lower()

        score = 0.0

        # 1. Platform match: does the URL actually contain the platform name?
        if platform.lower() in url:
            score += 0.4

        # 2. Name Match: How similar is the result title to the artist name?
        sim = name_similarity(artist_name, title)
        score += sim

        # 3. Official Signal: Does the title mention "official" or "artist"?
        if any(word in title for word in ["official", "artist", "music", "verified"]):
            score += 0.2

        # 4. Anti_signal: Penalty for common "wrong" pages
        if any(word in url for word in ["status", "video", "post", "event"]):
            score -= 0.3

        if score > max_score:
            max_score = score
            best_link = item.get("link")


    #Only return the link if it meets a minimum confidence threshold
    return best_link if max_score > 0.7 else "" 

def safe_sheet_name(name: str, used: set) -> str:
    """Cleans a string to be a vlid, unique Excel sheet name."""
    base = re.sub(INVALID_SHEET_CHARS, "_", (name or "Artist"))
    base = base[:MAX_SHEET_NAME_LEN].strip()

    candidate = base 
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = (base[:MAX_SHEET_NAME_LEN - len(suffix)] + suffix)
        i += 1
    used.add(candidate)
    return candidate

def main():
    parser = argparse.ArgumentParser(description="Artist Social Media Resolver")
    parser.add_argument("input", help="Input .xlsx file (artist_name, spotify_link, instagram_id)")
    parser.add_argument("output", help="Output .xlsx file to create")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"Error: {args.input} not found.")
        return

    try:
        # 1. Load and clean input data
        df = pd.read_excel(args.input)
        df.columns = [str(col).strip() for col in df.columns]

        if 'artist_name' in df.columns:
            df['artist_name'] = df['artist_name'].apply(normalize_whitespace)
        if 'spotify_link' in df.columns:
            df['spotify_link'] = df['spotify_link'].apply(normalize_spotify_link)
        if 'instagram_id' in df.columns:
            df['instagram_id'] = df['instagram_id'].apply(normalize_instagram_id)

        # 2. Initialize the Multi_sheet Workbook
        wb = Workbook()
        used_sheets = set()

        print(f"Processing {len(df)} artists...")

        # 3. The Master Loop
        for index, row in df.iterrows():
            artist_name = row.get('artist_name', 'Unknown Artist')
            spotify_link = row.get('spotify_link', '')
            insta_id = row.get('instagram_id', '')

            print(f"\n--- [{index + 1}/{len(df)}] {artist_name} ---")

            # Create a sheet for this artist
            sheet_title = safe_sheet_name(artist_name, used_sheets)
            ws = wb.create_sheet(title=sheet_title)
            ws.append(["Platform", "Correct Link", "Confidence"])

            # Step a: high confidence look up
            mb_identity = musicbrainz_lookup_artist(artist_name, spotify_link)
            mb_links = mb_identity.get("links", {})
            isni = mb_identity.get("isni", "")

            if isni:
                print(f" Verified via ISNI: {isni}")

            for platform in PLATFORMS:
                final_link = ""
                confidence = "not_found"

                # Priority 1: Direct Inputs
                if platform.lower() == "spotify":
                    final_link = spotify_link
                    confidence = "input_provided"
                elif platform.lower() == "instagram":
                    if insta_id:
                        final_link = f"https://instagram.com/{insta_id}/"
                        confidence = "input_provided"

                # Priority 2: MusicBrainz Expert Data
                if not final_link:
                    final_link = mb_links.get(platform.lower(), "")
                    if final_link:
                        confidence = "verified_musicbrainz"

                # Priority 3: Web Discovery (DuckDuckGo + Scoring)
                if not final_link:
                    print(f" Searching web for {platform}...")
                    query = f"{artist_name} {platform} official music"
                    search_results = search_duckduckgo_json(query)

                    # Use our 'Judge' function to pick the best candidate
                    discovered_link = pick_best_link(artist_name, platform, search_results)
                    if discovered_link:
                        final_link = discovered_link
                        confidence = "web_discovery_score"

                # Add row to artist sheet
                ws.append([platform, final_link, confidence])

                # Small pause for API etuquette
                time.sleep(0.3)

            # Mandatory 1-second pause per artist for MusicBrainz compliance
            time.sleep(1.0)

        # 4. Finalize file
        # Remove the default 'sheet' created by Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["Sheet"])

        wb.save(args.output)
        print(f"\nSuccess! Full workbook saved to: {args.output}")

    except Exception as e:
            print (f"FATAL ERROR: {e}")

if __name__ == "__main__":
    main()